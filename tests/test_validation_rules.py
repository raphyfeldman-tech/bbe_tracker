from __future__ import annotations
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.validation.rules import run_all_rules, Severity


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def test_missing_sheet_raises_error(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    del wb["Evidence"]
    wb.save(out)
    wb = load_workbook(out)

    findings = run_all_rules(wb)
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("Evidence" in f.message for f in errors)


def test_orphan_evidence_id_is_error(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    wb["Ownership"].append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 2.0,
        "2025-10-01", "EV-MISSING"
    ])
    wb.save(out)
    wb = load_workbook(out)

    findings = run_all_rules(wb)
    msgs = [f.message for f in findings if f.severity is Severity.ERROR]
    assert any("EV-MISSING" in m for m in msgs)


def test_valid_evidence_reference_passes(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    wb["Evidence"].append([
        "EV-0001", "Ownership", "Sample", "evidence/ownership/ev-0001.pdf",
        "2025-10-01", "r@x", "2025-10-01T10:00:00"
    ])
    wb["Ownership"].append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 2.0,
        "2025-10-01", "EV-0001"
    ])
    wb.save(out)
    wb = load_workbook(out)

    findings = run_all_rules(wb)
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert not any("EV-0001" in f.message for f in errors)
