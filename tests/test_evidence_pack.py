from __future__ import annotations
from pathlib import Path
import shutil
import zipfile
from openpyxl import load_workbook
from bee_tracker.evidence_pack import build_evidence_pack


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def _seed_with_evidence(tmp_path):
    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")

    (entity / "evidence" / "ownership").mkdir(parents=True)
    (entity / "evidence" / "ownership" / "ev-0001.pdf").write_bytes(b"%PDF-1.4 fake")
    (entity / "evidence" / "suppliers").mkdir(parents=True)
    (entity / "evidence" / "suppliers" / "ev-cert-1.pdf").write_bytes(b"%PDF-1.4 cert")

    wb = load_workbook(entity / "BEE_Tracker.xlsx")
    wb["Evidence"].append([
        "EV-0001", "Ownership", "Ownership cert",
        "evidence/ownership/ev-0001.pdf",
        "2025-10-01", "r@x", "2025-10-01T10:00:00",
    ])
    wb["Evidence"].append([
        "EV-CERT-1", "Suppliers", "Supplier 1 cert",
        "evidence/suppliers/ev-cert-1.pdf",
        "2025-10-01", "r@x", "2025-10-01T10:00:00",
    ])
    wb.save(entity / "BEE_Tracker.xlsx")
    return root


def test_build_evidence_pack_zips_workbook_and_evidence(tmp_path):
    root = _seed_with_evidence(tmp_path)
    out_zip = tmp_path / "evidence_pack.zip"
    build_evidence_pack(
        root=root, entity_name="sample", output_zip=out_zip,
    )
    assert out_zip.exists()
    with zipfile.ZipFile(out_zip) as z:
        names = z.namelist()
        assert any(n.endswith("BEE_Tracker.xlsx") for n in names)
        assert any("ev-0001.pdf" in n for n in names)
        assert any("ev-cert-1.pdf" in n for n in names)


def test_build_evidence_pack_skips_missing_files(tmp_path):
    root = _seed_with_evidence(tmp_path)
    wb_path = root / "entities" / "sample" / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)
    wb["Evidence"].append([
        "EV-MISSING", "Ownership", "Missing",
        "evidence/ownership/does-not-exist.pdf",
        "2025-10-01", "r@x", "2025-10-01T10:00:00",
    ])
    wb.save(wb_path)

    out_zip = tmp_path / "evidence_pack.zip"
    build_evidence_pack(
        root=root, entity_name="sample", output_zip=out_zip,
    )
    with zipfile.ZipFile(out_zip) as z:
        names = z.namelist()
        assert not any("does-not-exist.pdf" in n for n in names)
        assert any("ev-0001.pdf" in n for n in names)


def test_build_evidence_pack_includes_validation_report_when_present(tmp_path):
    root = _seed_with_evidence(tmp_path)
    val_html = root / "entities" / "sample" / "validation_latest.html"
    val_html.write_text("<html>validation</html>")

    out_zip = tmp_path / "evidence_pack.zip"
    build_evidence_pack(
        root=root, entity_name="sample", output_zip=out_zip,
    )
    with zipfile.ZipFile(out_zip) as z:
        names = z.namelist()
        assert any("validation_latest.html" in n for n in names)
