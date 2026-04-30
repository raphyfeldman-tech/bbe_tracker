from __future__ import annotations
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.cli.run_queue_daemon import process_one_entity


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
SCORECARD = Path("tests/fixtures/sample_ict_scorecard.yaml")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def _seed(tmp_path: Path, queue_rows: list) -> Path:
    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")
    wb = load_workbook(entity / "BEE_Tracker.xlsx")
    wb["Ownership"].append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 2.0,
        "2025-10-01", "EV-0001"
    ])
    for r in queue_rows:
        wb["RunQueue"].append(r)
    wb.save(entity / "BEE_Tracker.xlsx")
    return root


def test_daemon_processes_queued_score_request(tmp_path):
    root = _seed(tmp_path, [
        ["r-1", "2026-04-24T10:00:00", "u@x", "score",
         "queued", None, None, None, None],
    ])
    processed = process_one_entity(root=root, entity_name="sample")
    assert processed == 1

    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    ws = wb["RunQueue"]
    assert ws.cell(row=2, column=5).value == "completed"

    calc = wb["Calc_Ownership"]
    assert calc.max_row > 3


def test_daemon_skips_non_queued(tmp_path):
    root = _seed(tmp_path, [
        ["r-1", "2026-04-24T10:00:00", "u@x", "score",
         "completed", "2026-04-24T10:01:00", "2026-04-24T10:02:00", None, None],
    ])
    processed = process_one_entity(root=root, entity_name="sample")
    assert processed == 0


def test_daemon_marks_failed_on_bad_scope(tmp_path):
    root = _seed(tmp_path, [
        ["r-1", "2026-04-24T10:00:00", "u@x", "unknown_scope",
         "queued", None, None, None, None],
    ])
    processed = process_one_entity(root=root, entity_name="sample")
    assert processed == 1
    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    ws = wb["RunQueue"]
    assert ws.cell(row=2, column=5).value == "failed"
    assert "unknown_scope" in str(ws.cell(row=2, column=8).value)


def test_process_one_entity_accepts_explicit_backend(tmp_path):
    from openpyxl import load_workbook
    import shutil
    from bee_tracker.cli.run_queue_daemon import process_one_entity
    from bee_tracker.workbook.backends import LocalFolderBackend

    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")

    wb = load_workbook(entity / "BEE_Tracker.xlsx")
    wb["Ownership"].append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 2.0,
        "2025-10-01", "EV-0001"
    ])
    wb["RunQueue"].append([
        "r-explicit-backend", "2026-04-29T10:00:00", "u@x", "score",
        "queued", None, None, None, None,
    ])
    wb.save(entity / "BEE_Tracker.xlsx")

    backend = LocalFolderBackend(root)
    processed = process_one_entity(
        root=root, entity_name="sample", backend=backend,
    )
    assert processed == 1
    wb = load_workbook(entity / "BEE_Tracker.xlsx")
    assert wb["RunQueue"].cell(row=2, column=5).value == "completed"
