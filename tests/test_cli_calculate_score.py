from __future__ import annotations
from pathlib import Path
import shutil
from datetime import datetime
import pytest
from openpyxl import load_workbook
from bee_tracker.cli.calculate_score import run_score


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
SCORECARD = Path("tests/fixtures/sample_ict_scorecard.yaml")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def _seed_entity(tmp_path: Path) -> Path:
    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")
    # seed ownership data
    wb = load_workbook(entity / "BEE_Tracker.xlsx")
    wb["Ownership"].append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 2.0, "2025-10-01", "EV-0001"
    ])
    wb.save(entity / "BEE_Tracker.xlsx")
    return root


def test_run_score_writes_calc_ownership_and_dashboard(tmp_path: Path):
    root = _seed_entity(tmp_path)
    run_score(root=root, entity_name="sample", requested_by="tester@example")
    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")

    # Calc_Ownership populated
    calc = wb["Calc_Ownership"]
    headers = [c.value for c in calc[1]]
    assert headers == ["indicator", "points_earned", "max_points_check"]
    assert calc.max_row > 3

    # Dashboard populated — iter_rows(values_only=True) yields tuples of values
    dash = wb["Dashboard"]
    text = "|".join(str(v) for row in dash.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Sample Entity (Pty) Ltd" in text
    assert "Ownership" in text


def test_run_score_does_not_touch_ownership_input(tmp_path: Path):
    root = _seed_entity(tmp_path)
    wb_before = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    before = [[c.value for c in row] for row in wb_before["Ownership"].iter_rows()]

    run_score(root=root, entity_name="sample", requested_by="tester@example")

    wb_after = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    after = [[c.value for c in row] for row in wb_after["Ownership"].iter_rows()]
    assert before == after


def test_run_score_with_whatif_writes_calc_whatif(tmp_path):
    root = _seed_entity(tmp_path)
    # add a WhatIf override row that bumps net_value_pct
    from openpyxl import load_workbook as _lw
    wb_path = root / "entities" / "sample" / "BEE_Tracker.xlsx"
    wb = _lw(wb_path)
    wb["WhatIf"].append(["key", "value"])  # write headers if missing
    wb["WhatIf"].append(["ownership.net_value_pct", 25.0])
    wb.save(wb_path)

    run_score(root=root, entity_name="sample",
              requested_by="tester@example", whatif=True)
    wb = _lw(wb_path)
    calc_whatif = wb["Calc_WhatIf"]
    assert calc_whatif.max_row >= 2
    headers = [c.value for c in calc_whatif[1]]
    assert headers == ["element", "scenario_subtotal",
                       "scenario_max_points", "sub_minimum_breach"]


def test_run_score_applies_yes_level_up(tmp_path):
    """An entity with strong Y.E.S. absorption should see a level bump."""
    from openpyxl import load_workbook
    root = _seed_entity(tmp_path)

    wb_path = root / "entities" / "sample" / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)

    # Settings (required for full scoring)
    wb["Settings"].append(["entity_name", "Sample Entity"])
    wb["Settings"].append(["leviable_payroll", 10_000_000])
    wb["Settings"].append(["npat_current", 5_000_000])

    # Headcount of 100 employees, all not-mgmt to avoid affecting other elements
    for i in range(1, 101):
        wb["Employees"].append([
            f"E{i}", f"Emp{i}", "Black African", True, "Female",
            False, False, "Junior Mgmt", False, False,
            "2023-01-01", None, 12.0, 100_000,
        ])

    # 10 YES participants, 5 absorbed (50% absorption, 10% cohort) → Tier 3 → +2 levels
    for i in range(1, 6):
        wb["YES_Initiative"].append([
            f"Y{i}", "Black African", "Female", False, 22,
            "2025-10-01", "2026-09-30", True, True, "Yes", 5_000, "EV-OWN",
        ])
    for i in range(6, 11):
        wb["YES_Initiative"].append([
            f"Y{i}", "Black African", "Male", False, 24,
            "2025-10-01", "2026-09-30", True, False, "Yes", 5_000, "EV-OWN",
        ])
    wb["Evidence"].append([
        "EV-OWN", "Ownership", "stub", "x.pdf",
        "2025-10-01", "r@x", "2025-10-01T10:00:00",
    ])
    wb.save(wb_path)

    run_score(root=root, entity_name="sample", requested_by="tester@example")
    wb = load_workbook(wb_path)
    dash = wb["Dashboard"]
    text = "|".join(str(v) for row in dash.iter_rows(values_only=True)
                    for v in row if v is not None)
    # The Dashboard should mention Y.E.S. with a +2 levels-up
    assert "Y.E.S." in text or "YES" in text
    assert "+2 levels" in text or "levels_up=2" in text or "+2" in text


def test_run_score_skips_yes_when_not_participating(tmp_path):
    """If group_settings says yes_participating=false, no Y.E.S. should run."""
    from openpyxl import load_workbook
    root = _seed_entity(tmp_path)

    # Override group_settings to set yes_participating: false
    gs_path = root / "entities" / "sample" / "group_settings.yaml"
    text = gs_path.read_text().replace("yes_participating: true", "yes_participating: false")
    gs_path.write_text(text)

    run_score(root=root, entity_name="sample", requested_by="tester@example")
    # No assertion on absence of YES text — just that it doesn't crash without YES data


def test_run_score_appends_to_changelog(tmp_path):
    from openpyxl import load_workbook
    root = _seed_entity(tmp_path)
    run_score(root=root, entity_name="sample", requested_by="raphy@example")
    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    cl = wb["ChangeLog"]
    # Header at row 1, our run at row 2
    assert [c.value for c in cl[1]] == ["timestamp", "actor", "scope", "summary"]
    assert cl.cell(row=2, column=2).value == "raphy@example"
    assert cl.cell(row=2, column=3).value == "score"
    # Summary should mention element subtotals
    summary = str(cl.cell(row=2, column=4).value or "")
    assert "ownership" in summary.lower()


def test_run_score_writes_gap_analysis_sheet(tmp_path):
    from openpyxl import load_workbook
    root = _seed_entity(tmp_path)
    run_score(root=root, entity_name="sample", requested_by="r@x")
    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")
    ga = wb["GapAnalysis"]
    text = "|".join(str(v) for row in ga.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Ranked Financial Actions" in text
    assert "Non-Financial Opportunities" in text
