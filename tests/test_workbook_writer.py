from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import shutil
from bee_tracker.scoring.base import ElementResult
from bee_tracker.workbook.writer import write_calc_ownership


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def test_write_calc_ownership_populates_sheet(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)

    result = ElementResult(
        element="ownership",
        indicator_points={
            "black_voting_rights": 4.0,
            "net_value": 8.0,
        },
        subtotal=12.0,
        max_points=25.0,
        sub_minimum_breach=False,
    )
    write_calc_ownership(wb, result)
    wb.save(out)

    reopened = load_workbook(out)
    ws = reopened["Calc_Ownership"]
    # Row 1 = headers
    assert [c.value for c in ws[1]] == [
        "indicator", "points_earned", "max_points_check"
    ]
    # Subsequent rows = indicator points
    indicator_rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                      for r in range(2, ws.max_row + 1)
                      if ws.cell(row=r, column=1).value and ws.cell(row=r, column=1).value != "SUBTOTAL"}
    assert indicator_rows["black_voting_rights"] == 4.0
    assert indicator_rows["net_value"] == 8.0

    # Summary rows at bottom
    bottom_values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                     for r in range(2, ws.max_row + 1)}
    assert bottom_values["SUBTOTAL"] == 12.0
    assert bottom_values["MAX_POINTS"] == 25.0
    assert bottom_values["SUB_MINIMUM_BREACH"] in (False, "False", 0)


def test_write_gap_analysis_renders_both_sections(tmp_path):
    from openpyxl import load_workbook
    import shutil
    from bee_tracker.workbook.writer import write_gap_analysis
    from bee_tracker.gap_analysis.financial import Action
    from bee_tracker.gap_analysis.non_financial import Opportunity

    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)

    ranked = [
        Action("Shift R840k from Level 8 to Level 1", "enterprise_supplier_dev",
               840000, 1.4, 600000.0, "recognition multiplier increases"),
        Action("Add R310k training", "skills_development",
               310000, 1.2, 258333.33, "increases training_spend_pct"),
    ]
    opps = [
        Opportunity("Appoint 1 black female middle manager",
                    "management_control", 0.8, "hiring required"),
        Opportunity("Raise net_value_pct from 20% to 25%",
                    "ownership", 1.6, "strategic equity"),
    ]
    write_gap_analysis(wb, ranked_actions=ranked, opportunities=opps)
    wb.save(out)

    reopened = load_workbook(out)
    ws = reopened["GapAnalysis"]
    text = "|".join(str(v) for row in ws.iter_rows(values_only=True)
                    for v in row if v is not None)
    # Section A: ranked financial actions
    assert "Ranked Financial Actions" in text
    assert "Shift R840k from Level 8 to Level 1" in text
    assert "Add R310k training" in text
    # Section B: non-financial opportunities
    assert "Non-Financial Opportunities" in text
    assert "Appoint 1 black female middle manager" in text
    assert "Raise net_value_pct from 20% to 25%" in text


def test_write_gap_analysis_handles_empty_sections(tmp_path):
    from openpyxl import load_workbook
    import shutil
    from bee_tracker.workbook.writer import write_gap_analysis

    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    write_gap_analysis(wb, ranked_actions=[], opportunities=[])
    wb.save(out)

    # Should not raise, sheet should still have the section headers
    reopened = load_workbook(out)
    ws = reopened["GapAnalysis"]
    text = "|".join(str(v) for row in ws.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Ranked Financial Actions" in text
    assert "Non-Financial Opportunities" in text
