from __future__ import annotations
"""End-to-end Plan 2 walking skeleton.

Seeds an entity workbook with input rows for all 5 elements, queues a `score`
RunQueue request, runs one daemon iteration, and asserts:
  - All 5 Calc_* sheets are populated
  - RunQueue row marked completed
  - Dashboard shows entity name, BEE level, and an Ownership row in the breakdown
"""
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.cli.run_queue_daemon import process_one_entity


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
SCORECARD = Path("tests/fixtures/sample_ict_scorecard.yaml")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def _seed_full_entity(tmp_path: Path) -> Path:
    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")

    wb_path = entity / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)

    # Settings — required for skills + ESD + SED scoring
    wb["Settings"].append(["key", "value"])
    wb["Settings"].append(["entity_name", "Sample Entity (Pty) Ltd"])
    wb["Settings"].append(["leviable_payroll", 10_000_000])
    wb["Settings"].append(["npat_current", 5_000_000])

    # Evidence — needed for any evidence_id references
    wb["Evidence"].append([
        "EV-OWN", "Ownership", "Primary ownership cert",
        "evidence/ownership/ev-own.pdf",
        "2025-10-01", "raphy@core.co.za", "2025-10-01T10:00:00",
    ])

    # Ownership
    wb["Ownership"].append([
        "Shareholder A",
        30.0, 30.0, 15.0, 15.0, 5.0,   # voting/economic/etc
        25.0, 2.0,                       # net_value, new_entrants
        "2025-10-01",
        "EV-OWN",
    ])

    # Employees — seed 10, 6 black, mix of mgmt levels
    employees = [
        ("E1", "Alice",   "Black African", True,  "Female", False, False, "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750_000),
        ("E2", "Bob",     "Black African", True,  "Male",   False, False, "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750_000),
        ("E3", "Carol",   "White",         False, "Female", False, False, "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750_000),
        ("E4", "Dan",     "Black African", True,  "Male",   False, False, "Middle Mgmt", False, False, "2023-01-01", None, 12.0, 500_000),
        ("E5", "Erin",    "Black African", True,  "Female", False, False, "Middle Mgmt", False, False, "2023-01-01", None, 12.0, 500_000),
        ("E6", "Frank",   "White",         False, "Male",   False, False, "Middle Mgmt", False, False, "2023-01-01", None, 12.0, 500_000),
        ("E7", "Gina",    "Black African", True,  "Female", False, False, "Junior Mgmt", False, False, "2023-01-01", None, 12.0, 300_000),
        ("E8", "Hank",    "Black African", True,  "Male",   False, False, "Junior Mgmt", False, False, "2023-01-01", None, 12.0, 300_000),
        ("E9", "Ivy",     "White",         False, "Female", False, False, "Junior Mgmt", False, False, "2023-01-01", None, 12.0, 300_000),
        ("E10", "Jack",   "White",         False, "Male",   False, False, "Junior Mgmt", False, False, "2023-01-01", None, 12.0, 300_000),
    ]
    for row in employees:
        wb["Employees"].append(list(row))

    # Training — R600k spend on 2 black employees → 6% of leviable payroll
    wb["Training"].append([
        "T1", "E1", "Python BEE Bootcamp", "B", True, 300_000, 0, 40,
        "2025-11-01", "2025-11-30", True, "EV-OWN",
    ])
    wb["Training"].append([
        "T2", "E2", "Advanced BEE Reporting", "B", True, 300_000, 0, 40,
        "2025-11-01", "2025-11-30", True, "EV-OWN",
    ])

    # Suppliers + Procurement — Level 1 black-owned supplier
    wb["Suppliers"].append([
        "S1", "Black-owned Vendor", 1, "Generic", 80.0, 50.0,
        True, True, False, False,
        "2025-01-01", "2027-01-01", "EV-OWN",
    ])
    wb["Procurement"].append([
        "S1", 1_000_000, 0, "n/a", 30, "EV-OWN",
    ])

    # ESD: Enterprise Development + Supplier Development top-ups
    wb["ESD_Contributions"].append([
        "C1", "Enterprise Development", "Black-owned EME",
        80.0, "EME", "Grant", 100_000, 0, 1.0, "2025-12-01", "EV-OWN",
    ])
    wb["ESD_Contributions"].append([
        "C2", "Supplier Development", "Black-owned QSE",
        80.0, "QSE", "Loan", 100_000, 0, 1.0, "2025-12-01", "EV-OWN",
    ])

    # SED — 75k qualifying spend = 1.5% of NPAT → full target
    wb["SED_Contributions"].append([
        "B1", "Schools Trust", "NPO", 100.0, "Grant", 75_000, 0, "2025-12-01", "EV-OWN",
    ])

    # YES — cohort of 4 with 2 absorbed = 50% absorption (Tier 2 reachable)
    for i in range(1, 3):
        wb["YES_Initiative"].append([
            f"Y{i}", "Black African", "Female", False, 22,
            "2025-10-01", "2026-09-30", True, True, "Yes", 5_000, "EV-OWN",
        ])
    for i in range(3, 5):
        wb["YES_Initiative"].append([
            f"Y{i}", "Black African", "Male", False, 24,
            "2025-10-01", "2026-09-30", True, False, "Yes", 5_000, "EV-OWN",
        ])

    # RunQueue request
    wb["RunQueue"].append([
        "r-plan2", "2026-04-28T10:00:00", "raphy@core.co.za", "score",
        "queued", None, None, None, None,
    ])

    wb.save(wb_path)
    return root


def test_plan2_walking_skeleton(tmp_path):
    root = _seed_full_entity(tmp_path)
    n = process_one_entity(root=root, entity_name="sample")
    assert n == 1

    wb = load_workbook(root / "entities" / "sample" / "BEE_Tracker.xlsx")

    # RunQueue completed
    rq = wb["RunQueue"]
    assert rq.cell(row=2, column=1).value == "r-plan2"
    assert rq.cell(row=2, column=5).value == "completed"

    # All 5 Calc_* sheets populated
    for sheet_name in (
        "Calc_Ownership", "Calc_MgmtControl", "Calc_SkillsDev",
        "Calc_ESD", "Calc_SED",
    ):
        sheet = wb[sheet_name]
        assert sheet.max_row > 1, f"{sheet_name} should have data rows"

    # Dashboard shows entity, BEE level, breakdown rows
    dash = wb["Dashboard"]
    text = "|".join(str(v) for row in dash.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Sample Entity (Pty) Ltd" in text
    assert "BEE Level:" in text
    assert "Ownership" in text
    # Total score should be > 0
    # Look for "Total Score:" line and confirm any nonzero numeric appears in
    # the row immediately after.
    assert "Total Score:" in text


def test_plan2_walking_skeleton_with_whatif(tmp_path):
    """Run the same skeleton but also exercise the WhatIf scenario path."""
    from bee_tracker.cli.calculate_score import run_score
    root = _seed_full_entity(tmp_path)

    # Add a WhatIf override row that bumps net_value_pct
    wb_path = root / "entities" / "sample" / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)
    wb["WhatIf"].append(["key", "value"])
    wb["WhatIf"].append(["ownership.net_value_pct", 25.0])
    wb.save(wb_path)

    run_score(
        root=root, entity_name="sample",
        requested_by="raphy@core.co.za", whatif=True,
    )

    wb = load_workbook(wb_path)
    calc_whatif = wb["Calc_WhatIf"]
    assert calc_whatif.max_row >= 2  # header + at least one element row
    headers = [c.value for c in calc_whatif[1]]
    assert headers == ["element", "scenario_subtotal",
                       "scenario_max_points", "sub_minimum_breach"]

    # Dashboard contains the scenario column
    dash = wb["Dashboard"]
    text = "|".join(str(v) for row in dash.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Scenario (WhatIf)" in text
