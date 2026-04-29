from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import pytest
from bee_tracker.workbook.reader import read_ownership
from bee_tracker.workbook.reader import (
    read_employees, read_training, read_learnerships, read_bursaries,
    read_suppliers, read_procurement, read_esd_contributions,
    read_sed_contributions, read_yes_initiative, read_whatif,
    read_settings,
)


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def _seed_ownership(tmp_path: Path) -> Path:
    import shutil
    out = tmp_path / "BEE_Tracker.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    ws = wb["Ownership"]
    ws.append([
        "Shareholder A", 30.0, 30.0, 15.0, 15.0, 5.0, 25.0, 0.0, "2025-10-01", "EV-0001"
    ])
    ws.append([
        "Shareholder B", 10.0, 10.0,  0.0,  0.0, 0.0, 10.0, 0.0, "2025-10-01", "EV-0001"
    ])
    wb.save(out)
    return out


def test_read_ownership_returns_dataframe(tmp_path: Path):
    path = _seed_ownership(tmp_path)
    wb = load_workbook(path)
    df = read_ownership(wb)
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 2
    assert list(df.columns) == [
        "shareholder_name",
        "black_voting_rights_pct",
        "black_economic_interest_pct",
        "black_women_voting_rights_pct",
        "black_women_economic_interest_pct",
        "black_designated_groups_pct",
        "net_value_pct",
        "new_entrants_pct",
        "effective_date",
        "evidence_id",
    ]
    assert df.iloc[0]["shareholder_name"] == "Shareholder A"
    assert df.iloc[0]["black_voting_rights_pct"] == 30.0


def test_read_ownership_empty_returns_empty_df():
    wb = load_workbook(FIXTURE)
    df = read_ownership(wb)
    assert df.empty
    assert "shareholder_name" in df.columns


def test_read_employees_returns_dataframe(tmp_path):
    import shutil
    from openpyxl import load_workbook
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    wb["Employees"].append([
        "E1", "Alice", "Black African", True, "Female", False, False,
        "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750000,
    ])
    wb.save(out)
    df = read_employees(load_workbook(out))
    assert len(df) == 1
    assert df.iloc[0]["full_name"] == "Alice"


def test_read_settings_returns_dict(tmp_path):
    import shutil
    from openpyxl import load_workbook
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    wb["Settings"].append(["entity_name", "Sample Entity"])
    wb["Settings"].append(["leviable_payroll", 24000000])
    wb["Settings"].append(["npat_current", 15000000])
    wb.save(out)
    settings = read_settings(load_workbook(out))
    assert settings["entity_name"] == "Sample Entity"
    assert settings["leviable_payroll"] == 24000000
    assert settings["npat_current"] == 15000000


def test_read_suppliers_returns_dataframe(tmp_path):
    import shutil
    from openpyxl import load_workbook
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    wb["Suppliers"].append([
        "S1", "Acme Pty", 1, "Generic", 51.0, 30.0,
        True, True, False, False, "2025-01-01", "2026-01-01", "EV-CERT-1",
    ])
    wb.save(out)
    df = read_suppliers(load_workbook(out))
    assert len(df) == 1
    assert df.iloc[0]["cert_level"] == 1


def test_empty_sheets_return_dataframes_with_headers():
    from openpyxl import load_workbook
    wb = load_workbook(FIXTURE)
    assert "supplier_id" in read_suppliers(wb).columns
    assert "supplier_id" in read_procurement(wb).columns
    assert "participant_id" in read_yes_initiative(wb).columns
