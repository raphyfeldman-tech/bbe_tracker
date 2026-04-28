from __future__ import annotations
from pathlib import Path
import shutil
import pytest
from openpyxl import load_workbook, Workbook
from bee_tracker.workbook.backends import (
    LocalFolderBackend, WorkbookHandle,
)


@pytest.fixture
def local_root(tmp_path: Path) -> Path:
    entity = tmp_path / "entities" / "sample"
    entity.mkdir(parents=True)
    # seed with a minimal workbook
    wb = Workbook()
    wb.active.title = "Settings"
    wb.save(entity / "BEE_Tracker.xlsx")
    return tmp_path


def test_local_backend_opens_workbook(local_root: Path):
    backend = LocalFolderBackend(local_root)
    handle = backend.open_entity_workbook("sample")
    assert isinstance(handle, WorkbookHandle)
    assert "Settings" in handle.workbook.sheetnames


def test_local_backend_save_roundtrip_preserves_changes(local_root: Path):
    backend = LocalFolderBackend(local_root)
    handle = backend.open_entity_workbook("sample")
    handle.workbook["Settings"]["A1"] = "entity_name"
    handle.workbook["Settings"]["B1"] = "Sample Entity"
    backend.save(handle)

    reopened = load_workbook(local_root / "entities" / "sample" / "BEE_Tracker.xlsx")
    assert reopened["Settings"]["A1"].value == "entity_name"
    assert reopened["Settings"]["B1"].value == "Sample Entity"


def test_local_backend_missing_entity_raises(local_root: Path):
    backend = LocalFolderBackend(local_root)
    with pytest.raises(FileNotFoundError):
        backend.open_entity_workbook("does_not_exist")
