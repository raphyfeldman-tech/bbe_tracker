from __future__ import annotations
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.cli.run_queue_daemon import process_all_entities


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
SCORECARD = Path("tests/fixtures/sample_ict_scorecard.yaml")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def _seed_entity(root: Path, name: str, queued: bool) -> None:
    entity = root / "entities" / name
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    if queued:
        wb = load_workbook(entity / "BEE_Tracker.xlsx")
        wb["RunQueue"].append([
            f"r-{name}", "2026-04-28T10:00:00", "u@x", "score",
            "queued", None, None, None, None,
        ])
        wb.save(entity / "BEE_Tracker.xlsx")


def test_daemon_processes_all_entities(tmp_path):
    root = tmp_path / "bee_tracker"
    root.mkdir()
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")
    _seed_entity(root, "alpha", queued=True)
    _seed_entity(root, "beta",  queued=True)
    _seed_entity(root, "gamma", queued=False)

    total = process_all_entities(root=root)
    assert total == 2

    for name in ("alpha", "beta"):
        wb = load_workbook(root / "entities" / name / "BEE_Tracker.xlsx")
        assert wb["RunQueue"].cell(row=2, column=5).value == "completed"

    # Gamma had no queued requests → still has empty RunQueue (no row 2)
    wb_gamma = load_workbook(root / "entities" / "gamma" / "BEE_Tracker.xlsx")
    assert wb_gamma["RunQueue"].cell(row=2, column=5).value is None


def test_daemon_handles_missing_entities_dir(tmp_path):
    root = tmp_path / "bee_tracker"
    root.mkdir()
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")
    # No /entities/ at all
    total = process_all_entities(root=root)
    assert total == 0


def test_daemon_skips_non_directory_children(tmp_path):
    root = tmp_path / "bee_tracker"
    (root / "entities").mkdir(parents=True)
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")
    # Place a stray file under entities/ — should be skipped
    (root / "entities" / ".DS_Store").write_text("")
    _seed_entity(root, "alpha", queued=True)
    total = process_all_entities(root=root)
    assert total == 1
