from __future__ import annotations
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.cli.run_queue_daemon import process_one_entity


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")
SCORECARD = Path("tests/fixtures/sample_ict_scorecard.yaml")
GROUP_SETTINGS = Path("tests/fixtures/sample_group_settings.yaml")


def test_walking_skeleton(tmp_path):
    root = tmp_path / "bee_tracker"
    entity = root / "entities" / "sample"
    entity.mkdir(parents=True)
    shutil.copy(FIXTURE, entity / "BEE_Tracker.xlsx")
    shutil.copy(GROUP_SETTINGS, entity / "group_settings.yaml")
    shutil.copy(SCORECARD, root / "ict_scorecard.yaml")

    wb_path = entity / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)

    # Seed Evidence + Ownership + a queued request
    wb["Evidence"].append([
        "EV-0001", "Ownership", "Primary ownership certificate",
        "evidence/ownership/ev-0001.pdf",
        "2025-10-01", "raphy@core.co.za", "2025-10-01T10:00:00",
    ])
    wb["Ownership"].append([
        "Shareholder A",
        30.0,   # black_voting_rights_pct
        30.0,   # black_economic_interest_pct
        15.0,   # black_women_voting_rights_pct
        15.0,   # black_women_economic_interest_pct
        5.0,    # black_designated_groups_pct (target 3 → capped 3 pts)
        25.0,   # net_value_pct
        2.0,    # new_entrants_pct
        "2025-10-01",
        "EV-0001",
    ])
    wb["RunQueue"].append([
        "r-e2e-1", "2026-04-24T10:00:00", "raphy@core.co.za", "score",
        "queued", None, None, None, None,
    ])
    wb.save(wb_path)

    # One daemon iteration
    n = process_one_entity(root=root, entity_name="sample")
    assert n == 1

    # Assertions
    wb = load_workbook(wb_path)

    # RunQueue row marked completed
    rq = wb["RunQueue"]
    assert rq.cell(row=2, column=1).value == "r-e2e-1"
    assert rq.cell(row=2, column=5).value == "completed"

    # Calc_Ownership populated
    calc = wb["Calc_Ownership"]
    rows = {calc.cell(row=r, column=1).value: calc.cell(row=r, column=2).value
            for r in range(2, calc.max_row + 1)
            if calc.cell(row=r, column=1).value}
    assert rows["SUBTOTAL"] == 25.0
    assert rows["SUB_MINIMUM_BREACH"] in (False, 0, "False")

    # Dashboard shows entity + score — iter_rows(values_only=True) yields tuples of values
    dash = wb["Dashboard"]
    joined = "|".join(str(v) for row in dash.iter_rows(values_only=True)
                      for v in row if v is not None)
    assert "Sample Entity (Pty) Ltd" in joined
    assert "Ownership" in joined
    assert "25.0" in joined or "25" in joined
