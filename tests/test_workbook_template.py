from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE = Path("templates/workbook_template.xlsx")

EXPECTED_SHEETS = [
    "Dashboard", "GapAnalysis", "WhatIf",
    "Ownership", "Employees", "MgmtControl_Summary",
    "Training", "Learnerships", "Bursaries",
    "Suppliers", "Procurement",
    "ESD_Contributions", "SED_Contributions", "YES_Initiative",
    "Evidence",
    "Calc_Ownership", "Calc_MgmtControl", "Calc_SkillsDev",
    "Calc_ESD", "Calc_SED", "Calc_WhatIf",
    "History", "Settings",
    "Ref_Scorecard", "Ref_RecognitionLevels", "ChangeLog",
    "RunQueue",
]

TAB_COLOURS = {
    "Dashboard": "00B050",       # green
    "Ownership": "4472C4",       # blue
    "Settings": "BFBFBF",        # grey
    "RunQueue": "C00000",        # red
}


def test_template_exists():
    assert TEMPLATE.exists(), "Run scripts/make_template.py to regenerate."


def test_template_has_all_27_sheets_in_order():
    wb = load_workbook(TEMPLATE)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_template_tab_colours():
    wb = load_workbook(TEMPLATE)
    for sheet, colour in TAB_COLOURS.items():
        # openpyxl stores tab colour as an aRGB hex string (with alpha)
        actual = wb[sheet].sheet_properties.tabColor
        assert actual is not None, f"No tab colour set on {sheet}"
        assert colour in str(actual.value), f"{sheet} tab colour mismatch"


def test_ownership_sheet_has_expected_headers():
    wb = load_workbook(TEMPLATE)
    ws = wb["Ownership"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "shareholder_name",
        "black_voting_rights_pct",
        "black_economic_interest_pct",
        "black_women_voting_rights_pct",
        "black_women_economic_interest_pct",
        "black_designated_groups_pct",
        "net_value_pct",
        "new_entrants_pct",
        "effective_date",
        "evidence_id",
    ]


def test_runqueue_sheet_has_expected_headers():
    wb = load_workbook(TEMPLATE)
    ws = wb["RunQueue"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "request_id", "requested_at", "requested_by", "scope",
        "status", "started_at", "completed_at",
        "error_message", "result_path",
    ]


def test_template_is_byte_deterministic(tmp_path):
    """Regenerating the template twice must produce identical bytes.

    Without pinned timestamps in `build_template`, openpyxl stamps the current
    wall-clock time into docProps/core.xml and the committed binary becomes
    noisy and meaningless under git diff.
    """
    import sys
    scripts_dir = Path("scripts").resolve()
    sys.path.insert(0, str(scripts_dir))
    try:
        from make_template import build_template
    finally:
        sys.path.pop(0)

    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    build_template(a)
    build_template(b)
    assert a.read_bytes() == b.read_bytes()


def test_fixture_matches_committed_template():
    """The fixture copy must be byte-identical to the committed template."""
    assert (Path("tests/fixtures/sample_workbook.xlsx").read_bytes()
            == Path("templates/workbook_template.xlsx").read_bytes())
