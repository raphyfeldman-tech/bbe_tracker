from __future__ import annotations
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import shutil
from bee_tracker.scoring.base import ElementResult
from bee_tracker.rendering.dashboard import render_dashboard, DashboardContext


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def test_dashboard_shows_ownership_score(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)

    ctx = DashboardContext(
        entity_name="Sample Entity (Pty) Ltd",
        measurement_period="2025-10-01 → 2026-09-30",
        last_run_at=datetime(2026, 4, 24, 12, 0, 0),
        last_run_by="raphael.feldman@core.co.za",
        element_results=[
            ElementResult(
                element="ownership",
                indicator_points={},
                subtotal=18.5,
                max_points=25.0,
                sub_minimum_breach=False,
            ),
        ],
    )
    render_dashboard(wb, ctx)
    wb.save(out)

    reopened = load_workbook(out)
    ws = reopened["Dashboard"]
    text = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    joined = "|".join(text)
    assert "Sample Entity (Pty) Ltd" in joined
    assert "2025-10-01" in joined
    assert "Ownership" in joined
    assert "18.5" in joined
    assert "raphael.feldman@core.co.za" in joined


def test_dashboard_flags_priority_breach(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)

    ctx = DashboardContext(
        entity_name="Sample",
        measurement_period="2025-10-01 → 2026-09-30",
        last_run_at=datetime(2026, 4, 24),
        last_run_by="r@x",
        element_results=[
            ElementResult(
                element="ownership",
                indicator_points={},
                subtotal=5.0,
                max_points=25.0,
                sub_minimum_breach=True,
            ),
        ],
    )
    render_dashboard(wb, ctx)
    wb.save(out)

    reopened = load_workbook(out)
    ws = reopened["Dashboard"]
    text = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    joined = "|".join(text)
    assert "BREACH" in joined or "breach" in joined.lower()


def test_dashboard_renders_scenario_column(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)

    baseline = ElementResult(
        element="ownership", indicator_points={}, subtotal=10.0,
        max_points=25.0, sub_minimum_breach=False,
    )
    scenario = ElementResult(
        element="ownership", indicator_points={}, subtotal=20.0,
        max_points=25.0, sub_minimum_breach=False,
    )
    ctx = DashboardContext(
        entity_name="Sample",
        measurement_period="2025-10-01 → 2026-09-30",
        last_run_at=datetime(2026, 4, 28),
        last_run_by="r@x",
        element_results=[baseline],
        scenario_element_results=[scenario],
    )
    render_dashboard(wb, ctx)
    wb.save(out)

    reopened = load_workbook(out)
    ws = reopened["Dashboard"]
    text = "|".join(str(v) for row in ws.iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Scenario (WhatIf)" in text
    assert "20.0" in text or "20" in text  # scenario subtotal
    assert "10.0" in text or "10" in text  # delta = 10.0


def test_dashboard_renders_bee_level(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    ctx = DashboardContext(
        entity_name="Sample",
        measurement_period="2025-10-01 → 2026-09-30",
        last_run_at=datetime(2026, 4, 28),
        last_run_by="r@x",
        element_results=[
            ElementResult(element="ownership", indicator_points={},
                          subtotal=10.0, max_points=25.0, sub_minimum_breach=False),
        ],
        bee_level=5,
    )
    render_dashboard(wb, ctx)
    wb.save(out)

    reopened = load_workbook(out)
    text = "|".join(str(v) for row in reopened["Dashboard"].iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "BEE Level:" in text
    assert "5" in text


def test_dashboard_renders_top_gaps(tmp_path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    wb = load_workbook(out)
    ctx = DashboardContext(
        entity_name="Sample",
        measurement_period="2025-10-01 → 2026-09-30",
        last_run_at=datetime(2026, 4, 28),
        last_run_by="r@x",
        element_results=[
            ElementResult(element="ownership", indicator_points={},
                          subtotal=10.0, max_points=25.0, sub_minimum_breach=False),
        ],
        bee_level=5,
        top_gaps=[
            {"description": "Shift R840k to Level 1", "element": "enterprise_supplier_dev",
             "rand_required": 840000, "points_gained": 1.4, "rand_per_point": 600000},
            {"description": "Add R310k training", "element": "skills_development",
             "rand_required": 310000, "points_gained": 1.2, "rand_per_point": 258333},
        ],
    )
    render_dashboard(wb, ctx)
    wb.save(out)

    reopened = load_workbook(out)
    text = "|".join(str(v) for row in reopened["Dashboard"].iter_rows(values_only=True)
                    for v in row if v is not None)
    assert "Top Gaps to Next Level" in text
    assert "Shift R840k to Level 1" in text
    assert "Add R310k training" in text
