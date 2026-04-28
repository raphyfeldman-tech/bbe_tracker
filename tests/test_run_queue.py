from __future__ import annotations
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import shutil
import pytest
from bee_tracker.run_queue import (
    read_queued, mark_running, mark_completed, mark_failed, RunRequest,
)


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def _seed(path: Path, rows: list) -> None:
    wb = load_workbook(path)
    ws = wb["RunQueue"]
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_read_queued_returns_only_queued_rows(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    _seed(out, [
        ["r-1", "2026-04-24T10:00:00", "user@x", "score", "queued", None, None, None, None],
        ["r-2", "2026-04-24T10:05:00", "user@x", "score", "completed", None, None, None, None],
        ["r-3", "2026-04-24T10:10:00", "user@x", "score", "queued", None, None, None, None],
    ])
    wb = load_workbook(out)
    queued = read_queued(wb)
    assert [r.request_id for r in queued] == ["r-1", "r-3"]
    assert queued[0].scope == "score"


def test_mark_running_updates_status_and_timestamp(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    _seed(out, [
        ["r-1", "2026-04-24T10:00:00", "u", "score", "queued", None, None, None, None],
    ])
    wb = load_workbook(out)
    mark_running(wb, "r-1", started_at=datetime(2026, 4, 24, 10, 1, 0))
    wb.save(out)
    reopened = load_workbook(out)
    ws = reopened["RunQueue"]
    # status col = 5, started_at = 6
    assert ws.cell(row=2, column=5).value == "running"
    assert "2026-04-24T10:01:00" in str(ws.cell(row=2, column=6).value)


def test_mark_completed_updates_status_and_result(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    _seed(out, [
        ["r-1", "2026-04-24T10:00:00", "u", "score", "running",
         "2026-04-24T10:01:00", None, None, None],
    ])
    wb = load_workbook(out)
    mark_completed(wb, "r-1",
                   completed_at=datetime(2026, 4, 24, 10, 2, 0),
                   result_path="n/a")
    wb.save(out)
    reopened = load_workbook(out)
    ws = reopened["RunQueue"]
    assert ws.cell(row=2, column=5).value == "completed"
    assert "2026-04-24T10:02:00" in str(ws.cell(row=2, column=7).value)


def test_mark_failed_records_error(tmp_path: Path):
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    _seed(out, [
        ["r-1", "2026-04-24T10:00:00", "u", "score", "running",
         "2026-04-24T10:01:00", None, None, None],
    ])
    wb = load_workbook(out)
    mark_failed(wb, "r-1",
                completed_at=datetime(2026, 4, 24, 10, 2, 0),
                error_message="boom")
    wb.save(out)
    reopened = load_workbook(out)
    ws = reopened["RunQueue"]
    assert ws.cell(row=2, column=5).value == "failed"
    assert ws.cell(row=2, column=8).value == "boom"
