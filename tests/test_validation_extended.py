from __future__ import annotations
from datetime import date, datetime, timedelta
from pathlib import Path
import shutil
from openpyxl import load_workbook
from bee_tracker.validation.rules import run_all_rules, Severity


FIXTURE = Path("tests/fixtures/sample_workbook.xlsx")


def _seed(tmp_path: Path) -> Path:
    out = tmp_path / "wb.xlsx"
    shutil.copy(FIXTURE, out)
    return out


def test_supplier_cert_expired_is_error(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    wb["Suppliers"].append([
        "S1", "Old Cert Sup", 1, "Generic", 51.0, 30.0,
        True, True, False, False,
        "2024-01-01", "2025-01-01",  # expired
        "EV-CERT",
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("expired" in f.message.lower() and "S1" in f.message for f in errors)


def test_supplier_cert_expiring_30_days_is_warning(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    soon = (date.today() + timedelta(days=20)).isoformat()
    wb["Suppliers"].append([
        "S2", "Expiring Sup", 1, "Generic", 51.0, 30.0,
        True, True, False, False,
        "2024-01-01", soon, "EV-CERT-2",
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    warnings = [f for f in findings if f.severity is Severity.WARNING]
    assert any("S2" in f.message and ("30 days" in f.message or "expiring" in f.message.lower())
               for f in warnings)


def test_employee_missing_demographics_is_error(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    wb["Employees"].append([
        "E1", "Alice", None, True, "Female", False, False,    # missing race
        "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750000,
    ])
    wb["Employees"].append([
        "E2", "Bob", "Black African", True, None, False, False,    # missing gender
        "Senior Mgmt", False, False, "2023-01-01", None, 12.0, 750000,
    ])
    wb["Employees"].append([
        "E3", "Carol", "Black African", True, "Female", False, False,
        None, False, False, "2023-01-01", None, 12.0, 750000,    # missing level
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("E1" in f.message and "race" in f.message.lower() for f in errors)
    assert any("E2" in f.message and "gender" in f.message.lower() for f in errors)
    assert any("E3" in f.message and "occupational_level" in f.message.lower() for f in errors)


def test_settings_missing_npat_blocks_scoring(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    # Settings sheet has no npat_current/leviable_payroll rows
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("npat_current" in f.message for f in errors)
    assert any("leviable_payroll" in f.message for f in errors)


def test_esd_recipient_below_51pct_black_is_error(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    wb["ESD_Contributions"].append([
        "C1", "Enterprise Development", "Mostly White Co", 30.0,
        "QSE", "Grant", 100000, 0, 1.0, "2025-10-15", "EV-1",
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("C1" in f.message and "51" in f.message for f in errors)


def test_sed_beneficiary_below_75pct_black_is_warning(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    wb["SED_Contributions"].append([
        "B1", "Mixed Trust", "NPO", 60.0,
        "Grant", 50000, 0, "2025-10-15", "EV-1",
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    warnings = [f for f in findings if f.severity is Severity.WARNING]
    assert any("B1" in f.message and "75" in f.message for f in warnings)


def test_yes_participant_age_outside_range_is_error(tmp_path):
    out = _seed(tmp_path)
    wb = load_workbook(out)
    wb["YES_Initiative"].append([
        "Y1", "Black African", "Female", False, 17,
        "2025-10-01", "2026-09-30", True, False, "Yes", 5000, "EV-1",
    ])
    wb["YES_Initiative"].append([
        "Y2", "Black African", "Male", False, 36,
        "2025-10-01", "2026-09-30", True, False, "Yes", 5000, "EV-2",
    ])
    wb.save(out)
    findings = run_all_rules(load_workbook(out))
    errors = [f for f in findings if f.severity is Severity.ERROR]
    assert any("Y1" in f.message and "age" in f.message.lower() for f in errors)
    assert any("Y2" in f.message and "age" in f.message.lower() for f in errors)
