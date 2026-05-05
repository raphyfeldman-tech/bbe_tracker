from __future__ import annotations
"""Regenerate templates/workbook_template.xlsx from scratch.

Run: python scripts/make_template.py
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from bee_tracker.workbook.schema import SHEETS, TAB_COLOURS, HIDDEN_SHEETS


def _write_sheet(ws: Worksheet, tab_colour_key: str, headers: list[str]) -> None:
    ws.sheet_properties.tabColor = TAB_COLOURS[tab_colour_key]
    if headers:
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )
        ws.freeze_panes = "A2"
    if ws.title in HIDDEN_SHEETS:
        ws.sheet_state = "hidden"


def build_template(out: Path) -> None:
    wb = Workbook()
    # Pin timestamps so successive regenerations are byte-identical.
    # Without this, openpyxl stamps wall-clock time into docProps/core.xml
    # and the committed binary churns every time anyone runs this script.
    pinned = datetime(2026, 1, 1)
    wb.properties.created = pinned
    wb.properties.modified = pinned
    # openpyxl creates a default sheet; delete it
    default = wb.active
    wb.remove(default)
    for name, colour, headers in SHEETS:
        ws = wb.create_sheet(name)
        _write_sheet(ws, colour, headers)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    # Post-process: strip openpyxl's re-stamped modified timestamp.
    _strip_dc_modified(out)


def _strip_dc_modified(path: Path) -> None:
    """Replace dcterms:modified in docProps/core.xml with a pinned value
    and normalize per-entry zip timestamps.

    openpyxl re-stamps the modified-time during save even when we set
    wb.properties.modified beforehand, and the zip writer stamps each
    archive entry with the current wall-clock time. Post-processing the
    saved archive is the only reliable way to make the output
    byte-deterministic.
    """
    import re
    import shutil
    import zipfile
    tmp = path.with_suffix(".tmp")
    pinned = b"<dcterms:modified xsi:type=\"dcterms:W3CDTF\">2026-01-01T00:00:00Z</dcterms:modified>"
    pinned_dt = (2026, 1, 1, 0, 0, 0)
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for entry in zin.infolist():
            data = zin.read(entry.filename)
            if entry.filename == "docProps/core.xml":
                data = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                    pinned,
                    data,
                )
            # Pin per-entry mtime so the central directory is identical
            # across runs (zipfile stamps these from wall clock by default).
            entry.date_time = pinned_dt
            zout.writestr(entry, data)
    shutil.move(str(tmp), str(path))


if __name__ == "__main__":
    target = Path("templates/workbook_template.xlsx")
    build_template(target)
    print(f"Wrote {target}")
