from __future__ import annotations
"""Evidence-pack export — zip workbook + every referenced evidence file."""
import logging
import zipfile
from pathlib import Path
from openpyxl import load_workbook


log = logging.getLogger("bee_tracker.evidence_pack")


def build_evidence_pack(
    *,
    root: Path,
    entity_name: str,
    output_zip: Path,
) -> None:
    """Build a zip of the workbook + every referenced evidence file.

    Missing evidence files (referenced in the Evidence sheet but absent on
    disk) are skipped with a warning. Includes ``validation_latest.html``
    next to the workbook if present.
    """
    entity_dir = root / "entities" / entity_name
    wb_path = entity_dir / "BEE_Tracker.xlsx"
    output_zip.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(wb_path)
    if "Evidence" not in wb.sheetnames:
        evidence_paths: list[str] = []
    else:
        ws = wb["Evidence"]
        headers = [c.value for c in ws[1]]
        try:
            path_col = headers.index("filepath") + 1
        except ValueError:
            path_col = None
        evidence_paths = []
        if path_col is not None:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=path_col).value
                if v:
                    evidence_paths.append(str(v))

    with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(wb_path, arcname="BEE_Tracker.xlsx")
        for rel in evidence_paths:
            src = entity_dir / rel
            if not src.exists():
                log.warning("Evidence file missing on disk: %s", src)
                continue
            zf.write(src, arcname=rel)
        latest_validation = entity_dir / "validation_latest.html"
        if latest_validation.exists():
            zf.write(latest_validation, arcname="validation_latest.html")
