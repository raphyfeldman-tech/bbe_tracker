from __future__ import annotations
from openpyxl import Workbook
from ..scoring.base import ElementResult


def _clear_sheet(wb: Workbook, sheet: str) -> None:
    ws = wb[sheet]
    # openpyxl doesn't have a clear(); delete all rows then re-append headers
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def write_calc_element(wb: Workbook, sheet_name: str, result: ElementResult) -> None:
    """Generic Calc_<Element> writer.

    Use directly with the canonical sheet name, e.g.
        write_calc_element(wb, "Calc_Ownership", result)
    Or via the per-element shims (write_calc_ownership, etc.) for readability.
    """
    _clear_sheet(wb, sheet_name)
    ws = wb[sheet_name]
    ws.append(["indicator", "points_earned", "max_points_check"])
    for indicator, points in result.indicator_points.items():
        ws.append([indicator, points, None])
    ws.append([])
    ws.append(["SUBTOTAL", result.subtotal, None])
    ws.append(["MAX_POINTS", result.max_points, None])
    ws.append(["SUB_MINIMUM_BREACH", result.sub_minimum_breach, None])


def write_calc_ownership(wb: Workbook, result: ElementResult) -> None:
    """Overwrite Calc_Ownership with the given result. Does not touch any other sheet."""
    write_calc_element(wb, "Calc_Ownership", result)


def write_calc_mgmt_control(wb: Workbook, result: ElementResult) -> None:
    """Overwrite Calc_MgmtControl with the given result."""
    write_calc_element(wb, "Calc_MgmtControl", result)


def write_calc_skills_dev(wb: Workbook, result: ElementResult) -> None:
    """Overwrite Calc_SkillsDev with the given result."""
    write_calc_element(wb, "Calc_SkillsDev", result)


def write_calc_esd(wb: Workbook, result: ElementResult) -> None:
    """Overwrite Calc_ESD with the given result."""
    write_calc_element(wb, "Calc_ESD", result)


def write_calc_sed(wb: Workbook, result: ElementResult) -> None:
    """Overwrite Calc_SED with the given result."""
    write_calc_element(wb, "Calc_SED", result)


def write_calc_whatif(wb: Workbook, scenario_results: list[ElementResult]) -> None:
    """Overwrite Calc_WhatIf with scenario element subtotals."""
    _clear_sheet(wb, "Calc_WhatIf")
    ws = wb["Calc_WhatIf"]
    ws.append(["element", "scenario_subtotal", "scenario_max_points", "sub_minimum_breach"])
    for r in scenario_results:
        ws.append([r.element, r.subtotal, r.max_points, r.sub_minimum_breach])


def write_gap_analysis(
    wb: Workbook,
    *,
    ranked_actions: list,        # list[Action] from gap_analysis.financial
    opportunities: list,         # list[Opportunity] from gap_analysis.non_financial
) -> None:
    """Overwrite the GapAnalysis sheet.

    Section A: ranked financial actions (description | element | R required |
    points gained | R/point | reason).

    Section B: non-financial opportunities (description | element | points
    gained | notes).
    """
    _clear_sheet(wb, "GapAnalysis")
    ws = wb["GapAnalysis"]

    # Section A — Ranked financial actions
    ws.append(["Ranked Financial Actions"])
    ws.append([
        "Action", "Element", "R Required", "Points", "R / Point", "Reason",
    ])
    for a in ranked_actions:
        ws.append([
            a.description, a.element, a.rand_required,
            a.points_gained, a.rand_per_point, a.reason,
        ])
    ws.append([])

    # Section B — Non-financial opportunities
    ws.append(["Non-Financial Opportunities"])
    ws.append(["Opportunity", "Element", "Points Gained", "Notes"])
    for o in opportunities:
        ws.append([
            o.description, o.element, o.points_gained, o.notes,
        ])


def append_change_log(wb: Workbook, *, actor: str, scope: str, summary: str,
                      timestamp_iso: str) -> None:
    """Append a row to the ChangeLog sheet. Idempotent: ensures the header
    row exists, then appends below it."""
    ws = wb["ChangeLog"]
    expected_headers = ["timestamp", "actor", "scope", "summary"]
    current_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if current_headers != expected_headers:
        # Sheet is empty or has wrong header — clear and re-seed
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(expected_headers)
    ws.append([timestamp_iso, actor, scope, summary])
