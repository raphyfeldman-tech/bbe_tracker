from __future__ import annotations
"""Pluggable workbook backends.

GraphBackend is used in production against SharePoint.
LocalFolderBackend is used in tests and for local dev — reads/writes
directly from a folder that mirrors the SharePoint tree.
"""
from abc import ABC, abstractmethod
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook
from ..graph.client import GraphClient, ItemMetadata
from ..errors import WorkbookError


@dataclass
class WorkbookHandle:
    """In-memory representation of an entity's workbook during a run."""
    entity_name: str
    workbook: Workbook
    # Opaque token used by the backend to detect concurrent modifications.
    # GraphBackend puts the eTag here; LocalFolderBackend puts the mtime.
    original_token: str | None = None


class WorkbookBackend(ABC):
    @abstractmethod
    def open_entity_workbook(self, entity_name: str) -> WorkbookHandle: ...

    @abstractmethod
    def save(self, handle: WorkbookHandle) -> None: ...


class LocalFolderBackend(WorkbookBackend):
    def __init__(self, root: Path):
        self._root = Path(root)

    def _workbook_path(self, entity_name: str) -> Path:
        return self._root / "entities" / entity_name / "BEE_Tracker.xlsx"

    def open_entity_workbook(self, entity_name: str) -> WorkbookHandle:
        path = self._workbook_path(entity_name)
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        wb = load_workbook(path)
        return WorkbookHandle(
            entity_name=entity_name,
            workbook=wb,
            original_token=str(path.stat().st_mtime_ns),
        )

    def save(self, handle: WorkbookHandle) -> None:
        path = self._workbook_path(handle.entity_name)
        handle.workbook.save(path)


class GraphBackend(WorkbookBackend):
    """Reads and writes workbooks on SharePoint via Microsoft Graph.

    `entity_locator` maps entity names to (drive_id, item_id) pairs.
    """

    def __init__(self, client: GraphClient, entity_locator: dict[str, tuple[str, str]]):
        self._client = client
        self._locator = entity_locator

    def _resolve(self, entity_name: str) -> tuple[str, str]:
        if entity_name not in self._locator:
            raise WorkbookError(f"Unknown entity: {entity_name}")
        return self._locator[entity_name]

    def open_entity_workbook(self, entity_name: str) -> WorkbookHandle:
        drive_id, item_id = self._resolve(entity_name)
        data, meta = self._client.download_item(drive_id, item_id)
        wb = load_workbook(BytesIO(data))
        return WorkbookHandle(
            entity_name=entity_name,
            workbook=wb,
            original_token=meta.etag,
        )

    def save(self, handle: WorkbookHandle) -> None:
        drive_id, item_id = self._resolve(handle.entity_name)
        buf = BytesIO()
        handle.workbook.save(buf)
        meta = self._client.upload_item(
            drive_id, item_id, buf.getvalue(),
            if_match=handle.original_token or "*",
        )
        # Refresh the token so subsequent saves of the same handle don't
        # fail with a 412 against the previous version we just wrote.
        handle.original_token = meta.etag

    @classmethod
    def from_env(cls, locator_yaml_path: Path) -> "GraphBackend":
        """Build a GraphBackend from environment variables + a locator YAML.

        Required env vars:
          GRAPH_TENANT_ID, GRAPH_CLIENT_ID, GRAPH_CLIENT_SECRET

        Locator YAML format:
          entity_a:
            drive_id: "<drive id>"
            item_id: "<item id of BEE_Tracker.xlsx>"
          entity_b:
            ...
        """
        import os
        import yaml
        from ..graph.auth import GraphAuth
        from ..graph.client import GraphClient

        auth = GraphAuth(
            tenant_id=os.environ["GRAPH_TENANT_ID"],
            client_id=os.environ["GRAPH_CLIENT_ID"],
            client_secret=os.environ["GRAPH_CLIENT_SECRET"],
        )
        client = GraphClient(auth)
        locator_data = yaml.safe_load(Path(locator_yaml_path).read_text()) or {}
        entity_locator = {
            name: (cfg["drive_id"], cfg["item_id"])
            for name, cfg in locator_data.items()
        }
        return cls(client, entity_locator)
