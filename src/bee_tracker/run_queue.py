from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from openpyxl import Workbook


# RunQueue column layout (1-indexed):
COL_REQUEST_ID       = 1
COL_REQUESTED_AT     = 2
COL_REQUESTED_BY     = 3
COL_SCOPE            = 4
COL_STATUS           = 5
COL_STARTED_AT       = 6
COL_COMPLETED_AT     = 7
COL_ERROR_MESSAGE    = 8
COL_RESULT_PATH      = 9

VALID_SCOPES = {"score", "score_with_whatif", "report_monthly",
                "report_quarterly", "evidence_pack"}


@dataclass(frozen=True)
class RunRequest:
    request_id: str
    requested_at: str
    requested_by: str
    scope: str


def _iter_data_rows(wb: Workbook):
    ws = wb["RunQueue"]
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[COL_REQUEST_ID - 1].value is None:
            continue
        yield idx, row


def _find_row(wb: Workbook, request_id: str) -> int:
    for idx, row in _iter_data_rows(wb):
        if row[COL_REQUEST_ID - 1].value == request_id:
            return idx
    raise KeyError(f"RunQueue row not found: {request_id}")


def read_queued(wb: Workbook) -> list[RunRequest]:
    out: list[RunRequest] = []
    for _, row in _iter_data_rows(wb):
        if row[COL_STATUS - 1].value == "queued":
            out.append(RunRequest(
                request_id=row[COL_REQUEST_ID - 1].value,
                requested_at=str(row[COL_REQUESTED_AT - 1].value),
                requested_by=row[COL_REQUESTED_BY - 1].value,
                scope=row[COL_SCOPE - 1].value,
            ))
    return out


def mark_running(wb: Workbook, request_id: str, *, started_at: datetime) -> None:
    row = _find_row(wb, request_id)
    ws = wb["RunQueue"]
    ws.cell(row=row, column=COL_STATUS, value="running")
    ws.cell(row=row, column=COL_STARTED_AT, value=started_at.isoformat())


def mark_completed(
    wb: Workbook, request_id: str, *, completed_at: datetime, result_path: str
) -> None:
    row = _find_row(wb, request_id)
    ws = wb["RunQueue"]
    ws.cell(row=row, column=COL_STATUS, value="completed")
    ws.cell(row=row, column=COL_COMPLETED_AT, value=completed_at.isoformat())
    ws.cell(row=row, column=COL_RESULT_PATH, value=result_path)


def mark_failed(
    wb: Workbook, request_id: str, *, completed_at: datetime, error_message: str
) -> None:
    row = _find_row(wb, request_id)
    ws = wb["RunQueue"]
    ws.cell(row=row, column=COL_STATUS, value="failed")
    ws.cell(row=row, column=COL_COMPLETED_AT, value=completed_at.isoformat())
    ws.cell(row=row, column=COL_ERROR_MESSAGE, value=error_message)
