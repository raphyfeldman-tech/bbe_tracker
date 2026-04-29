from __future__ import annotations
"""Polls each entity's RunQueue and dispatches jobs.

Plan 1 scope: one entity at a time, 'score' scope only.
Plan 2 expands: multiple entities, full scope coverage (reports, evidence pack).
"""
import argparse
import logging
import time
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from ..run_queue import read_queued, mark_running, mark_completed, mark_failed
from .calculate_score import run_score


log = logging.getLogger("bee_tracker.run_queue_daemon")


def process_one_entity(*, root: Path, entity_name: str) -> int:
    """Process all queued rows for a single entity. Returns count processed."""
    wb_path = root / "entities" / entity_name / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)
    queued = read_queued(wb)
    if not queued:
        return 0

    processed = 0
    for req in queued:
        mark_running(wb, req.request_id, started_at=datetime.utcnow())
        wb.save(wb_path)

        try:
            if req.scope == "score":
                run_score(
                    root=root,
                    entity_name=entity_name,
                    requested_by=req.requested_by,
                )
                # re-open after run_score so we see the latest RunQueue
                wb = load_workbook(wb_path)
                mark_completed(
                    wb, req.request_id,
                    completed_at=datetime.utcnow(),
                    result_path="(in-workbook)",
                )
            else:
                raise ValueError(f"Unsupported scope in Plan 1: {req.scope}")
        except Exception as exc:
            wb = load_workbook(wb_path)
            mark_failed(
                wb, req.request_id,
                completed_at=datetime.utcnow(),
                error_message=str(exc),
            )
            log.exception("Request %s failed", req.request_id)

        wb.save(wb_path)
        processed += 1

    return processed


def process_all_entities(*, root: Path) -> int:
    """Iterate every directory under root/entities/ and call process_one_entity."""
    entities_root = root / "entities"
    if not entities_root.exists():
        return 0
    total = 0
    for child in sorted(entities_root.iterdir()):
        if not child.is_dir():
            continue
        try:
            total += process_one_entity(root=root, entity_name=child.name)
        except Exception:
            log.exception("Entity %s failed; continuing", child.name)
    return total


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="bee-run-queue-daemon")
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--entity", default=None,
                        help="Process only this entity (default: process all entities under root/entities/)")
    parser.add_argument("--interval", type=int, default=60,
                        help="Seconds between polls")
    parser.add_argument("--once", action="store_true",
                        help="Run one iteration and exit")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    while True:
        try:
            if args.entity:
                n = process_one_entity(root=args.root, entity_name=args.entity)
            else:
                n = process_all_entities(root=args.root)
            if n:
                log.info("Processed %d request(s)", n)
        except Exception:
            log.exception("Iteration failed; continuing")
        if args.once:
            return 0
        time.sleep(args.interval)


if __name__ == "__main__":
    raise SystemExit(main())
