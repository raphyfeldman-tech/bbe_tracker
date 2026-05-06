from __future__ import annotations
"""``bee-send-alerts`` entry point — fan out alerts via Graph."""
import argparse
import logging
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from ..config import load_group_settings, load_scorecard
from ..workbook.reader import (
    read_ownership, read_employees, read_training, read_learnerships,
    read_bursaries, read_suppliers, read_procurement,
    read_esd_contributions, read_sed_contributions, read_settings,
)
from ..scoring.registry import default_registry
from ..scoring.level import level_after_priority_breaches
from ..alerts.triggers import (
    detect_priority_breaches, detect_cert_expiries, detect_level_drop,
)
from ..alerts.render import (
    render_priority_breach, render_cert_expiry, render_level_drop,
)
from ..alerts.email import send_email
from ..graph.auth import GraphAuth
from ..graph.client import GraphClient


log = logging.getLogger("bee_tracker.send_alerts")


def run_send_alerts(
    *,
    root: Path,
    entity_name: str,
    graph_client: GraphClient,
    from_user: str,
    templates_dir: Path,
    today: date | None = None,
) -> None:
    today = today or date.today()
    scorecard = load_scorecard(root / "ict_scorecard.yaml")
    gs = load_group_settings(
        root / "entities" / entity_name / "group_settings.yaml"
    )
    wb_path = root / "entities" / entity_name / "BEE_Tracker.xlsx"
    wb = load_workbook(wb_path)

    inputs = {
        "ownership": read_ownership(wb),
        "employees": read_employees(wb),
        "training": read_training(wb),
        "learnerships": read_learnerships(wb),
        "bursaries": read_bursaries(wb),
        "suppliers": read_suppliers(wb),
        "procurement": read_procurement(wb),
        "esd_contributions": read_esd_contributions(wb),
        "sed_contributions": read_sed_contributions(wb),
        "settings": read_settings(wb),
    }
    results = [s.score(inputs, scorecard) for s in default_registry().values()]

    # 1. Priority breach
    breaches = detect_priority_breaches(results)
    breach_recipients = gs.alerts.get("priority_element_breach", [])
    if breaches and breach_recipients:
        body = render_priority_breach(
            entity_name=gs.entity_name, breaches=breaches,
            templates_dir=templates_dir,
        )
        send_email(
            graph_client, from_user=from_user, to=breach_recipients,
            subject=f"[BEE Tracker] Priority element BREACH for {gs.entity_name}",
            html_body=body,
        )
        log.info("Sent priority-breach alert: %d recipient(s)", len(breach_recipients))

    # 2. Cert expiry
    expiries = detect_cert_expiries(inputs["suppliers"], today=today)
    expiry_recipients = gs.alerts.get("cert_expiry", [])
    if expiries and expiry_recipients:
        body = render_cert_expiry(
            entity_name=gs.entity_name, expiries=expiries,
            templates_dir=templates_dir,
        )
        send_email(
            graph_client, from_user=from_user, to=expiry_recipients,
            subject=f"[BEE Tracker] Supplier certs expiring for {gs.entity_name}",
            html_body=body,
        )
        log.info("Sent cert-expiry alert: %d recipient(s)", len(expiry_recipients))

    # 3. Level drop — read prior_level from History sheet (row 2 col 1 by convention)
    drop_recipients = gs.alerts.get("level_drop", [])
    if drop_recipients:
        history_ws = wb["History"]
        prior_level = None
        if history_ws.max_row >= 2:
            prior_level = history_ws.cell(row=2, column=1).value
        total_score = round(sum(r.subtotal for r in results), 4)
        breach_count = len(breaches)
        current_level = level_after_priority_breaches(
            total_score, breach_count=breach_count, scorecard=scorecard,
        )
        if prior_level is not None and detect_level_drop(current_level, prior_level):
            body = render_level_drop(
                entity_name=gs.entity_name,
                prior_level=prior_level, current_level=current_level,
                templates_dir=templates_dir,
            )
            send_email(
                graph_client, from_user=from_user, to=drop_recipients,
                subject=f"[BEE Tracker] Level dropped for {gs.entity_name}",
                html_body=body,
            )
            log.info("Sent level-drop alert: %d recipient(s)", len(drop_recipients))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="bee-send-alerts")
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--entity", required=True)
    parser.add_argument("--from-user", required=True,
                        help="Mailbox to send alerts as")
    parser.add_argument("--templates", type=Path, default=Path("templates"),
                        dest="templates_dir")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    import os
    auth = GraphAuth(
        tenant_id=os.environ["GRAPH_TENANT_ID"],
        client_id=os.environ["GRAPH_CLIENT_ID"],
        client_secret=os.environ["GRAPH_CLIENT_SECRET"],
    )
    client = GraphClient(auth)

    run_send_alerts(
        root=args.root,
        entity_name=args.entity,
        graph_client=client,
        from_user=args.from_user,
        templates_dir=args.templates_dir,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
