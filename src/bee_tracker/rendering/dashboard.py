from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ..scoring.base import ElementResult


@dataclass(frozen=True)
class DashboardContext:
    entity_name: str
    measurement_period: str
    last_run_at: datetime
    last_run_by: str
    element_results: list[ElementResult]
    bee_level: int | str = "non_compliant"
    scenario_element_results: list[ElementResult] | None = None
    top_gaps: list[dict] | None = None
    yes_levels_up: int = 0


def _clear(wb: Workbook, sheet: str) -> None:
    ws = wb[sheet]
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def render_dashboard(wb: Workbook, ctx: DashboardContext) -> None:
    """Render the Dashboard. Overwrites; never touches input sheets."""
    _clear(wb, "Dashboard")
    ws = wb["Dashboard"]

    # Header tiles
    ws.append(["B-BBEE Dashboard"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Entity:", ctx.entity_name])
    ws.append(["Measurement Period:", ctx.measurement_period])
    ws.append([])

    # Total score (sum of element subtotals for now)
    total = sum(r.subtotal for r in ctx.element_results)
    ws.append(["Total Score:", total])
    ws.append([])

    # BEE level tile
    if ctx.yes_levels_up > 0:
        ws.append(["BEE Level:", str(ctx.bee_level), f"(Y.E.S. +{ctx.yes_levels_up} levels)"])
    else:
        ws.append(["BEE Level:", str(ctx.bee_level)])
    ws.append([])

    # Priority-element status strip
    ws.append(["Priority Element Status"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for r in ctx.element_results:
        status = "BREACH" if r.sub_minimum_breach else "OK"
        row = [r.element.title(), status]
        ws.append(row)
        if r.sub_minimum_breach:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
            )
    ws.append([])

    # Element breakdown
    ws.append(["Element Breakdown"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Element", "Subtotal", "Max Points"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    for r in ctx.element_results:
        ws.append([r.element.title(), r.subtotal, r.max_points])
    ws.append([])

    # Scenario column when WhatIf scenario was run
    if ctx.scenario_element_results:
        ws.append(["Scenario (WhatIf)"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Element", "Scenario Subtotal", "Δ vs Baseline"])
        for header_col in (1, 2, 3):
            ws.cell(row=ws.max_row, column=header_col).font = Font(bold=True)
        baseline_by_name = {r.element: r.subtotal for r in ctx.element_results}
        for r in ctx.scenario_element_results:
            baseline = baseline_by_name.get(r.element, 0.0)
            delta = round(r.subtotal - baseline, 4)
            ws.append([r.element.title(), r.subtotal, delta])
        ws.append([])

    # Top 5 gaps
    if ctx.top_gaps:
        ws.append(["Top Gaps to Next Level"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Action", "Element", "R Required", "Points", "R / Point"])
        for header_col in range(1, 6):
            ws.cell(row=ws.max_row, column=header_col).font = Font(bold=True)
        for gap in ctx.top_gaps[:5]:
            ws.append([
                gap["description"], gap["element"],
                gap["rand_required"], gap["points_gained"], gap["rand_per_point"],
            ])
        ws.append([])

    # Run record
    ws.append(["Last recalc:", ctx.last_run_at.isoformat(), "by", ctx.last_run_by])
