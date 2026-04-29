from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from enum import Enum
from openpyxl import Workbook
import pandas as pd

from ..workbook.schema import EXPECTED_SHEET_NAMES as EXPECTED_SHEETS
from ..workbook.reader import (
    read_suppliers, read_employees, read_settings,
    read_esd_contributions, read_sed_contributions, read_yes_initiative,
)


class Severity(Enum):
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass(frozen=True)
class Finding:
    severity: Severity
    sheet: str | None
    row: int | None
    message: str

# Which sheets carry evidence_id references
EVIDENCE_REF_SHEETS = [
    "Ownership", "Training", "Learnerships", "Bursaries",
    "Procurement", "ESD_Contributions", "SED_Contributions",
    "YES_Initiative",
]


def check_structural(wb: Workbook) -> list[Finding]:
    missing = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
    return [
        Finding(Severity.ERROR, None, None, f"Missing expected sheet: {s}")
        for s in missing
    ]


def _collect_evidence_ids(wb: Workbook) -> set[str]:
    if "Evidence" not in wb.sheetnames:
        return set()
    ws = wb["Evidence"]
    headers = [c.value for c in ws[1]]
    if "evidence_id" not in headers:
        return set()
    col = headers.index("evidence_id") + 1
    return {ws.cell(row=r, column=col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=col).value}


def check_evidence_references(wb: Workbook) -> list[Finding]:
    known = _collect_evidence_ids(wb)
    findings: list[Finding] = []
    for sheet in EVIDENCE_REF_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        # Find evidence_id-shaped columns (may include cert_evidence_id too)
        cols = [(i + 1, h) for i, h in enumerate(headers)
                if h and str(h).endswith("evidence_id")]
        for col_idx, col_name in cols:
            for row_idx in range(2, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None or v == "":
                    continue
                if v not in known:
                    findings.append(Finding(
                        Severity.ERROR,
                        sheet,
                        row_idx,
                        f"Orphan {col_name}: {v!r} in {sheet}!{col_name} "
                        f"row {row_idx} (not present in Evidence sheet)",
                    ))
    return findings


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def check_supplier_cert_expiry(wb: Workbook, today: date | None = None) -> list[Finding]:
    today = today or date.today()
    findings: list[Finding] = []
    if "Suppliers" not in wb.sheetnames:
        return findings
    df = read_suppliers(wb)
    if df.empty or "supplier_id" not in df.columns:
        return findings
    for idx, row in df.iterrows():
        sid = row.get("supplier_id")
        expiry = _parse_date(row.get("cert_expiry_date"))
        if expiry is None:
            continue
        days = (expiry - today).days
        if days < 0:
            findings.append(Finding(
                Severity.ERROR, "Suppliers", idx + 2,
                f"Supplier {sid} cert expired on {expiry.isoformat()} "
                f"(before today {today.isoformat()})",
            ))
        elif days < 30:
            findings.append(Finding(
                Severity.WARNING, "Suppliers", idx + 2,
                f"Supplier {sid} cert expiring in {days} days "
                f"(< 30 days, on {expiry.isoformat()})",
            ))
        elif days < 60:
            findings.append(Finding(
                Severity.WARNING, "Suppliers", idx + 2,
                f"Supplier {sid} cert expiring in {days} days "
                f"(< 60 days, on {expiry.isoformat()})",
            ))
    return findings


_REQUIRED_EMPLOYEE_FIELDS = ("race", "gender", "occupational_level")


def check_employee_demographics(wb: Workbook) -> list[Finding]:
    findings: list[Finding] = []
    if "Employees" not in wb.sheetnames:
        return findings
    df = read_employees(wb)
    if df.empty or "employee_id" not in df.columns:
        return findings
    for idx, row in df.iterrows():
        eid = row.get("employee_id")
        for field_name in _REQUIRED_EMPLOYEE_FIELDS:
            value = row.get(field_name)
            if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
                findings.append(Finding(
                    Severity.ERROR, "Employees", idx + 2,
                    f"Employee {eid} missing {field_name}",
                ))
    return findings


def check_settings_required_fields(wb: Workbook) -> list[Finding]:
    findings: list[Finding] = []
    settings = read_settings(wb)
    for required in ("npat_current", "leviable_payroll"):
        v = settings.get(required)
        if v is None or v == "":
            findings.append(Finding(
                Severity.ERROR, "Settings", None,
                f"Settings missing required field: {required} "
                f"(blocks scoring)",
            ))
    return findings


def check_esd_recipient_threshold(wb: Workbook) -> list[Finding]:
    findings: list[Finding] = []
    if "ESD_Contributions" not in wb.sheetnames:
        return findings
    df = read_esd_contributions(wb)
    if df.empty:
        return findings
    for idx, row in df.iterrows():
        cid = row.get("contribution_id")
        pct = row.get("recipient_black_ownership_pct")
        if pct is None or pd.isna(pct):
            continue
        if float(pct) < 51:
            findings.append(Finding(
                Severity.ERROR, "ESD_Contributions", idx + 2,
                f"ESD contribution {cid} recipient is "
                f"{pct}% black-owned (< 51% threshold)",
            ))
    return findings


def check_sed_beneficiary_threshold(wb: Workbook) -> list[Finding]:
    findings: list[Finding] = []
    if "SED_Contributions" not in wb.sheetnames:
        return findings
    df = read_sed_contributions(wb)
    if df.empty:
        return findings
    for idx, row in df.iterrows():
        cid = row.get("contribution_id")
        pct = row.get("black_beneficiary_pct")
        if pct is None or pd.isna(pct):
            continue
        if float(pct) < 75:
            findings.append(Finding(
                Severity.WARNING, "SED_Contributions", idx + 2,
                f"SED contribution {cid} beneficiary is {pct}% black "
                f"(< 75% threshold; will not be recognised)",
            ))
    return findings


def check_yes_participant_age(wb: Workbook) -> list[Finding]:
    findings: list[Finding] = []
    if "YES_Initiative" not in wb.sheetnames:
        return findings
    df = read_yes_initiative(wb)
    if df.empty:
        return findings
    for idx, row in df.iterrows():
        pid = row.get("participant_id")
        age = row.get("age_at_start")
        if age is None or pd.isna(age):
            continue
        try:
            age_int = int(age)
        except (TypeError, ValueError):
            continue
        if age_int < 18 or age_int > 35:
            findings.append(Finding(
                Severity.ERROR, "YES_Initiative", idx + 2,
                f"YES participant {pid} age {age_int} is outside 18-35 range",
            ))
    return findings


def run_all_rules(wb: Workbook) -> list[Finding]:
    out: list[Finding] = []
    out.extend(check_structural(wb))
    out.extend(check_evidence_references(wb))
    out.extend(check_supplier_cert_expiry(wb))
    out.extend(check_employee_demographics(wb))
    out.extend(check_settings_required_fields(wb))
    out.extend(check_esd_recipient_threshold(wb))
    out.extend(check_sed_beneficiary_threshold(wb))
    out.extend(check_yes_participant_age(wb))
    return out
